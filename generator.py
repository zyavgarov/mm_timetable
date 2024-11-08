import json
from datetime import datetime

import numpy as np
import openpyxl as xl
import pandas as pd


def is_group_name(cell):
    # the function gets a celle value and returns if it is a group name
    try:
        if int(cell) >= 100:
            return True
    except ValueError:
        if cell[:2] in ['М1', 'М2'] or cell[:3] == 'ВСО':
            return True
    return False


def is_timetable(page_name):
    # the function gets a page of a workbook and checks if it is a timetable page
    try:
        if 0 < int(page_name[0]) < 7 and page_name[1] == '.':
            return True
        if int(page_name[1:3]) == 41:
            return True
    except ValueError:
        if page_name in ['М', 'ВСО']:
            return True
    return False


def are_times_far_enough(time1, time2):
    # the function gets two time moments and checks if the difference between them is more than 1 hour and 36 minutes
    date1 = datetime.combine(datetime.today().date(), time1)
    date2 = datetime.combine(datetime.today().date(), time2)
    if abs((date1 - date2).total_seconds()) > 5700:
        return True
    return False


def get_subject_name(cell):
    # the function gets a cell value and extracts subject name
    return str(cell).replace(':', ',').split(',', 1)[0]


def convert_to_name(group_name):
    # the function gets a group name and converts it to integer string. if the name is not a float then group name left untouched
    try:
        return str(int(group_name))
    except ValueError:
        return group_name


def str_to_time(time_string):
    # the function gets a time string and converts it to time object
    # accepted formats are 'HH:MM' and 'HH:MM:SS'
    if len(time_string) == 5:
        return datetime.strptime(time_string, '%H:%M').time()
    if len(time_string) == 8:
        return datetime.strptime(time_string, '%H:%M:%S').time()
    return time_string


def get_distribution(group_list, max_group, target_sum):
    # the function gets a list of groups and returns a distribution of groups
    # I assume a well-behaved user therefore I won't raise a warning in case of impossibility of the distribution
    distribution = np.random.rand(len(group_list))
    distribution = [int(x) for x in np.around(distribution * max_group).astype(int)]
    if np.sum(distribution) > target_sum:
        delta = np.sum(distribution) - target_sum
        for i in range(delta):
            k = np.random.randint(len(distribution))
            distribution[k] = distribution[k] - 1
    elif np.sum(distribution) < target_sum:
        delta = int(target_sum - np.sum(distribution))
        for i in range(delta):
            k = np.random.randint(len(distribution))
            distribution[k] = distribution[k] + 1
    return distribution


timetable_book = xl.load_workbook('timetable.xlsx')
lists = []
page_names = timetable_book.sheetnames
for i, sheet in enumerate(timetable_book):
    merged_ranges = sheet.merged_cells.ranges
    list_merged_ranges = list(merged_ranges)
    for merged_range in list_merged_ranges:
        start_cell = merged_range.start_cell
        cell_value = sheet[start_cell.coordinate].value
        sheet.unmerge_cells(range_string=str(merged_range))
        for row in merged_range.rows:
            #print('cv',row)
            for cell in row:
                sheet.cell(row=cell[0], column=cell[1]).value = cell_value
                #print(sheet.cell(row=cell[0], column=cell[1]).value)

for i, sheet in enumerate(timetable_book):
    data = sheet.values
    columns = next(data)
    lists.append(pd.DataFrame(data, columns=columns))

#print(lists[:5])
timetable_dict = {}
for page_num, page in enumerate(lists):

    # check if the page is a timetable
    if not is_timetable(page_names[page_num]):
        break

    # get groups
    current_row = 0
    while current_row < page.shape[0]:
        if not page.iloc[current_row, 2] is None:
            if is_group_name(page.iloc[current_row, 2]):
                break
        current_row += 1
    groups = []
    current_column = 2
    while current_column < page.shape[1] and page.iloc[current_row, current_column] is not None and is_group_name(
            page.iloc[current_row, current_column]):
        groups.append(convert_to_name(page.iloc[current_row, current_column]))
        current_column += 1

    days_of_week = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']
    times = page.iloc[:, 1]
    for group in groups:
        timetable_dict[group] = {}
    day_pointer = -1
    for row, time in enumerate(times):
        if pd.isnull(time) or str(time).split(' ')[0] in ['--', 'группа', 'кафедра', 'Расписание']:
            continue
        if isinstance(time, str):
            time = time.replace('-', ':')
            time = str_to_time(time)
        if isinstance(time, datetime):
            time = str(time)
            time = time.replace('-', ':')
            time = time[8:10] + ":" + time[5:7]
            time = str_to_time(time)
        if str(time) == "09:00:00" or (str(groups[0])[:3] == 'ВСО' and str(time) == "18:30:00"):
            day_pointer += 1
            for group in groups:
                timetable_dict[group][days_of_week[day_pointer]] = {}
        if not timetable_dict[groups[0]][days_of_week[day_pointer]] or are_times_far_enough(str_to_time((list(timetable_dict[groups[0]][days_of_week[day_pointer]].keys()))[-1]), time):
            for i, group in enumerate(groups):
                subj_name = get_subject_name(page.iloc[row, 2 + i])
                if subj_name == 'День самостоятельной работы' or subj_name.split(' ')[0] == 'ЕНС':
                    subj_name = 'None'
                timetable_dict[group][days_of_week[day_pointer]][str(time)] = subj_name

# saving to json
with open('timetable.json', 'w') as f:
    json.dump(timetable_dict, f, indent=4, ensure_ascii=False)

# I randomly distribute students by groups
# I assume that a group can't be larger than 30 people
with open('people_by_year.json', 'r') as f:
    distr_by_year = json.load(f)
groups_size = {}
for year in distr_by_year:
    group_list = [key for key in timetable_dict.keys() if key[0] == year]
    distr_by_group = get_distribution(group_list, 30, distr_by_year[year])
    for i, group_name in enumerate(group_list):
        groups_size[group_name] = distr_by_group[i]

with open('groups_size.json', 'w') as f:
    json.dump(groups_size, f, indent=4, ensure_ascii=False)